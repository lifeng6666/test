import os
import sys
import json
import glob
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 全局变量用于存储汇总数据
all_accounts = []
summary_info = {
    'total_groups': 0,
    'total_accounts': 0,
    'success_accounts': 0,
    'failed_accounts': [],
    'password_error_accounts': []
}

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def parse_result_file(filepath):
    """解析单个账号组的结果文件"""
    accounts = []
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            data = json.loads(content)
            
            group_index = data.get('group_index', 0)
            group_accounts = data.get('accounts', [])
            
            for acc in group_accounts:
                acc['group_index'] = group_index
                accounts.append(acc)
                
                # 更新汇总信息
                if acc.get('password_error', False):
                    summary_info['password_error_accounts'].append({
                        'index': acc.get('account_index', 0),
                        'group': group_index,
                        'username': acc.get('username', '')
                    })
                elif not acc.get('query_success', False):
                    summary_info['failed_accounts'].append({
                        'index': acc.get('account_index', 0),
                        'group': group_index,
                        'username': acc.get('username', '')
                    })
                else:
                    summary_info['success_accounts'] += 1
                    
            summary_info['total_accounts'] += len(group_accounts)
            summary_info['total_groups'] = max(summary_info['total_groups'], group_index)
            
    except Exception as e:
        log(f"解析文件 {filepath} 失败: {e}")
    
    return accounts

def collect_all_results():
    """收集所有账号组的结果文件"""
    global all_accounts
    
    # 查找所有结果文件
    result_files = glob.glob('jlc_result_*.json')
    
    if not result_files:
        log("未找到任何结果文件")
        return False
    
    log(f"找到 {len(result_files)} 个结果文件")
    
    for filepath in result_files:
        accounts = parse_result_file(filepath)
        all_accounts.extend(accounts)
        log(f"从 {filepath} 解析了 {len(accounts)} 个账号")
    
    # 按金豆数量排序（由高到低，失败的放后面）
    def sort_key(x):
        if x.get('query_success'):
            return int(x.get('jindou', 0))
        return -1
    
    all_accounts.sort(key=sort_key, reverse=True)
    
    return True

def generate_excel():
    """生成Excel排名文件"""
    # 获取当前日期
    now = datetime.now()
    month = now.month
    day = now.day
    total_accounts = summary_info['total_accounts']
    
    filename = f"{month}.{day}立创金豆查询(共查询{total_accounts}个账号).xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "金豆排名"
    
    # 设置标题行
    headers = ['排名', '金豆数量', '客编', '密码', '归属账号组']
    ws.append(headers)
    
    # 设置标题行样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    total_cols = len(headers)  # 5列
    
    # 先记录每个账号的全局排名，供后续 Sheet 使用
    for rank, acc in enumerate(all_accounts, 1):
        acc['global_rank'] = rank
    
    # ==============================
    # 填充第一个 Sheet（全局金豆排名）
    # ==============================
    for rank, acc in enumerate(all_accounts, 1):
        username = acc.get('username', '')
        actual_password = acc.get('actual_password', '')
        group_index = acc.get('group_index', 0)
        query_success = acc.get('query_success', False)
        password_error = acc.get('password_error', False)
        
        display_password = actual_password if actual_password else ''
        
        # 将异常状态展示在金豆数量列中
        if password_error:
            display_jindou = '密码错误'
        elif not query_success:
            display_jindou = acc.get('query_status', '查询失败')
        else:
            display_jindou = acc.get('jindou', 0)
        
        row_data = [
            rank,
            display_jindou,
            username,
            display_password,
            f"{group_index}组账号{acc.get('account_index', 0)}"
        ]
        ws.append(row_data)
        
        # 设置数据行样式
        row_num = rank + 1
        for col_num in range(1, total_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 排名列加粗
            if col_num == 1:
                cell.font = Font(bold=True)
            
            # 金豆数量列使用不同颜色区分
            if col_num == 2:
                if isinstance(display_jindou, (int, float)):
                    if display_jindou >= 500:
                        cell.font = Font(color="C00000", bold=True)  # 深红色
                    elif display_jindou >= 300:
                        cell.font = Font(color="FF6600", bold=True)  # 橙色
                    elif display_jindou >= 100:
                        cell.font = Font(color="0070C0", bold=True)  # 蓝色
                else:
                    # 异常文字直接标红
                    cell.font = Font(color="C00000", bold=True)
    
    # ==============================
    # 填充第二个 Sheet（按组及原顺序排序）
    # ==============================
    ws_group = wb.create_sheet("按组顺序")
    ws_group.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_group.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 按照组别和原本的账号顺序排序
    accounts_by_group = sorted(all_accounts, key=lambda x: (x.get('group_index', 0), x.get('account_index', 0)))
    
    current_row = 2
    last_group = None
    
    for acc in accounts_by_group:
        group_index = acc.get('group_index', 0)
        
        if last_group is not None and group_index != last_group:
            ws_group.append([''] * total_cols)
            for col_num in range(1, total_cols + 1):
                cell = ws_group.cell(row=current_row, column=col_num)
                cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
            current_row += 1
            
        last_group = group_index
        
        username = acc.get('username', '')
        actual_password = acc.get('actual_password', '')
        query_success = acc.get('query_success', False)
        password_error = acc.get('password_error', False)
        
        display_password = actual_password if actual_password else ''
        
        if password_error:
            display_jindou = '密码错误'
        elif not query_success:
            display_jindou = acc.get('query_status', '查询失败')
        else:
            display_jindou = acc.get('jindou', 0)
            
        global_rank = acc.get('global_rank', '')
        
        row_data = [
            global_rank, display_jindou, username, display_password,
            f"{group_index}组账号{acc.get('account_index', 0)}"
        ]
        ws_group.append(row_data)
        
        for col_num in range(1, total_cols + 1):
            cell = ws_group.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col_num == 1:
                cell.font = Font(bold=True)
            if col_num == 2:
                if isinstance(display_jindou, (int, float)):
                    if display_jindou >= 500:
                        cell.font = Font(color="C00000", bold=True)
                    elif display_jindou >= 300:
                        cell.font = Font(color="FF6600", bold=True)
                    elif display_jindou >= 100:
                        cell.font = Font(color="0070C0", bold=True)
                else:
                    cell.font = Font(color="C00000", bold=True)
                    
        current_row += 1

    # ==============================
    # 填充第三个 Sheet（按组排名 - 组内金豆从高到低）
    # ==============================
    ws_group_rank = wb.create_sheet("按组排名")
    ws_group_rank.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_group_rank.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 按照组别升序，再按金豆数量降序
    def rank_sort_key(x):
        g_index = x.get('group_index', 0)
        if x.get('query_success'):
            return (g_index, -int(x.get('jindou', 0)))
        return (g_index, 0)
        
    accounts_by_group_rank = sorted(all_accounts, key=rank_sort_key)
    
    current_row_rank = 2
    last_group_rank = None
    
    for acc in accounts_by_group_rank:
        group_index = acc.get('group_index', 0)
        
        if last_group_rank is not None and group_index != last_group_rank:
            ws_group_rank.append([''] * total_cols)
            for col_num in range(1, total_cols + 1):
                cell = ws_group_rank.cell(row=current_row_rank, column=col_num)
                cell.fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
            current_row_rank += 1
            
        last_group_rank = group_index
        
        username = acc.get('username', '')
        actual_password = acc.get('actual_password', '')
        query_success = acc.get('query_success', False)
        password_error = acc.get('password_error', False)
        
        display_password = actual_password if actual_password else ''
        
        if password_error:
            display_jindou = '密码错误'
        elif not query_success:
            display_jindou = acc.get('query_status', '查询失败')
        else:
            display_jindou = acc.get('jindou', 0)
            
        global_rank = acc.get('global_rank', '')
        
        row_data = [
            global_rank, display_jindou, username, display_password,
            f"{group_index}组账号{acc.get('account_index', 0)}"
        ]
        ws_group_rank.append(row_data)
        
        for col_num in range(1, total_cols + 1):
            cell = ws_group_rank.cell(row=current_row_rank, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if col_num == 1:
                cell.font = Font(bold=True)
            if col_num == 2:
                if isinstance(display_jindou, (int, float)):
                    if display_jindou >= 500:
                        cell.font = Font(color="C00000", bold=True)
                    elif display_jindou >= 300:
                        cell.font = Font(color="FF6600", bold=True)
                    elif display_jindou >= 100:
                        cell.font = Font(color="0070C0", bold=True)
                else:
                    cell.font = Font(color="C00000", bold=True)
                    
        current_row_rank += 1

    # ==============================
    # 全局格式设置（列宽、边框、冻结窗格）
    # ==============================
    column_widths = [8, 25, 18, 15, 18]  # 金豆数量列加宽以容纳错误文字
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 格式化 Sheet 1
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
    ws.freeze_panes = 'A2'
    
    # 格式化 Sheet 2
    for i, width in enumerate(column_widths, 1):
        ws_group.column_dimensions[get_column_letter(i)].width = width
    for row in ws_group.iter_rows(min_row=1, max_row=ws_group.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
    ws_group.freeze_panes = 'A2'

    # 格式化 Sheet 3
    for i, width in enumerate(column_widths, 1):
        ws_group_rank.column_dimensions[get_column_letter(i)].width = width
    for row in ws_group_rank.iter_rows(min_row=1, max_row=ws_group_rank.max_row, min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = thin_border
    ws_group_rank.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(filename)
    log(f"Excel文件已生成: {filename}")
    
    return filename

def get_push_title():
    """获取推送标题"""
    now = datetime.now()
    return f"{now.month}月{now.day}日立创金豆查询结果"

def get_push_content():
    """获取推送内容"""
    now = datetime.now()
    month = now.month
    day = now.day
    total_accounts = summary_info['total_accounts']
    
    failed_count = len(summary_info['failed_accounts'])
    pwd_error_count = len(summary_info['password_error_accounts'])
    
    if failed_count == 0 and pwd_error_count == 0:
        return f"{month}月{day}日立创金豆查询已全部成功(共查询{total_accounts}个账号)"
    else:
        content_parts = []
        
        if failed_count > 0:
            # 按组分类失败账号
            failed_by_group = {}
            for acc in summary_info['failed_accounts']:
                group = acc['group']
                if group not in failed_by_group:
                    failed_by_group[group] = []
                failed_by_group[group].append(str(acc['index']))
            
            for group, indices in failed_by_group.items():
                content_parts.append(f"{group}组账号{','.join(indices)}")
        
        if pwd_error_count > 0:
            pwd_by_group = {}
            for acc in summary_info['password_error_accounts']:
                group = acc['group']
                if group not in pwd_by_group:
                    pwd_by_group[group] = []
                pwd_by_group[group].append(str(acc['index']))
            
            for group, indices in pwd_by_group.items():
                content_parts.append(f"{group}组账号{','.join(indices)}(密码错误)")
        
        return f"{month}月{day}日立创金豆查询有{'/'.join(content_parts)}失败(共查询{total_accounts}个账号)"

def get_workflow_url():
    """获取GitHub Actions工作流运行页面链接"""
    server_url = os.getenv('GITHUB_SERVER_URL', 'https://github.com')
    repository = os.getenv('GITHUB_REPOSITORY', '')
    run_id = os.getenv('GITHUB_RUN_ID', '')
    
    if repository and run_id:
        return f"{server_url}/{repository}/actions/runs/{run_id}"
    return ""

def push_to_telegram(text, excel_file=None):
    """推送到Telegram"""
    bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
    chat_id = os.getenv('TELEGRAM_CHAT_ID')
    
    if not bot_token or not chat_id:
        return False
    
    try:
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        params = {'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'}
        response = requests.post(url, params=params, timeout=30)
        if response.status_code == 200:
            log("Telegram-文字消息已推送")
        else:
            return False
            
        if excel_file and os.path.exists(excel_file):
            doc_url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
            with open(excel_file, 'rb') as f:
                files = {'document': (os.path.basename(excel_file), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                requests.post(doc_url, data={'chat_id': chat_id}, files=files, timeout=60)
        return True
    except Exception:
        return False

def push_to_wechat(text, excel_file=None):
    """推送到企业微信"""
    webhook_key = os.getenv('WECHAT_WEBHOOK_KEY')
    
    if not webhook_key:
        return False
    
    try:
        url = webhook_key if webhook_key.startswith('https://') else f"https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={webhook_key}"
        body = {"msgtype": "text", "text": {"content": text}}
        requests.post(url, json=body, timeout=30)
        
        if excel_file and os.path.exists(excel_file):
            key = webhook_key.split('key=')[-1] if 'key=' in webhook_key else webhook_key
            upload_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file"
            with open(excel_file, 'rb') as f:
                files = {'media': (os.path.basename(excel_file), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                upload_response = requests.post(upload_url, files=files, timeout=60)
                if upload_response.status_code == 200:
                    media_id = upload_response.json().get('media_id')
                    send_url = f"https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}"
                    requests.post(send_url, json={"msgtype": "file", "file": {"media_id": media_id}}, timeout=30)
        return True
    except Exception:
        return False

def push_to_dingtalk(text):
    """推送到钉钉"""
    webhook = os.getenv('DINGTALK_WEBHOOK')
    if not webhook: return False
    try:
        url = webhook if webhook.startswith('https://') else f"https://oapi.dingtalk.com/robot/send?access_token={webhook}"
        requests.post(url, json={"msgtype": "text", "text": {"content": text}}, timeout=30)
        return True
    except Exception: return False

def push_to_pushplus(text):
    """推送到PushPlus"""
    token = os.getenv('PUSHPLUS_TOKEN')
    if not token: return False
    try:
        requests.post("http://www.pushplus.plus/send", json={"token": token, "title": get_push_title(), "content": text}, timeout=30)
        return True
    except Exception: return False

def push_to_serverchan(text):
    """推送到Server酱"""
    sckey = os.getenv('SERVERCHAN_SCKEY')
    if not sckey: return False
    try:
        requests.post(f"https://sctapi.ftqq.com/{sckey}.send", data={"title": get_push_title(), "desp": text}, timeout=30)
        return True
    except Exception: return False

def push_to_serverchan3(text):
    """推送到Server酱3"""
    sckey = os.getenv('SERVERCHAN3_SCKEY')
    if not sckey: return False
    try:
        from serverchan_sdk import sc_send
        sc_send(sckey, get_push_title(), text, {"tags": "嘉立创|查询"})
        return True
    except Exception: return False

def push_to_coolpush(text):
    """推送到酷推"""
    skey = os.getenv('COOLPUSH_SKEY')
    if not skey: return False
    try:
        requests.get(f"https://push.xuthus.cc/send/{skey}?c={text}", timeout=30)
        return True
    except Exception: return False

def push_to_custom(text):
    """推送到自定义API"""
    webhook = os.getenv('CUSTOM_WEBHOOK')
    if not webhook: return False
    try:
        requests.post(webhook, json={"title": get_push_title(), "content": text}, timeout=30)
        return True
    except Exception: return False

def push_all_notifications(excel_file):
    """推送所有通知"""
    title = get_push_title()
    content = get_push_content()
    workflow_url = get_workflow_url()
    
    push_text = f"{title}\n\n{content}"
    if workflow_url:
        push_text += f"\n请访问以下链接，在Artifacts板块下载金豆详细排名：\n{workflow_url}"
    
    log(f"推送内容:\n{push_text}")
    
    push_to_telegram(push_text, excel_file)
    push_to_wechat(push_text, excel_file)
    push_to_dingtalk(push_text)
    push_to_pushplus(push_text)
    push_to_serverchan(push_text)
    push_to_serverchan3(push_text)
    push_to_coolpush(push_text)
    push_to_custom(push_text)

def clean_temp_files():
    """清理临时结果文件"""
    temp_files = glob.glob('jlc_result_*.json')
    for filepath in temp_files:
        try: os.remove(filepath)
        except Exception: pass

def main():
    log("开始汇总金豆查询结果...")
    if not collect_all_results(): return
    if not all_accounts: return
    
    log(f"共收集到 {len(all_accounts)} 个账号的数据")
    excel_file = generate_excel()
    push_all_notifications(excel_file)
    clean_temp_files()
    log("汇总完成!")

if __name__ == "__main__":
    main()